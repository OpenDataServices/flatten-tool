"""
This file has classes describing input from spreadsheets.

"""

from __future__ import print_function, unicode_literals

import datetime
import os
from collections import OrderedDict, UserDict
from csv import DictReader
from csv import reader as csvreader
from decimal import Decimal, InvalidOperation
from warnings import warn

import openpyxl
import pytz
from openpyxl.utils.cell import _get_column_letter

from flattentool.exceptions import DataErrorWarning
from flattentool.i18n import _
from flattentool.lib import isint, parse_sheet_configuration
from flattentool.ODSReader import ODSReader

try:
    from zipfile import BadZipFile
except ImportError:
    from zipfile import BadZipfile as BadZipFile


class Cell:
    def __init__(self, cell_value, cell_location):
        self.cell_value = cell_value
        self.cell_location = cell_location
        self.sub_cells = []


def convert_type(type_string, value, timezone=pytz.timezone("UTC")):
    if value == "" or value is None:
        return None
    if type_string == "number":
        try:
            return Decimal(value)
        except (TypeError, ValueError, InvalidOperation):
            warn(
                _(
                    'Non-numeric value "{}" found in number column, returning as string instead.'
                ).format(value),
                DataErrorWarning,
            )
            return str(value)
    elif type_string == "integer":
        try:
            return int(value)
        except (TypeError, ValueError):
            warn(
                _(
                    'Non-integer value "{}" found in integer column, returning as string instead.'
                ).format(value),
                DataErrorWarning,
            )
            return str(value)
    elif type_string == "boolean":
        value = str(value)
        if value.lower() in ["true", "1"]:
            return True
        elif value.lower() in ["false", "0"]:
            return False
        else:
            warn(
                _(
                    'Unrecognised value for boolean: "{}", returning as string instead'
                ).format(value),
                DataErrorWarning,
            )
            return str(value)
    elif type_string in ("array", "array_array", "string_array", "number_array"):
        value = str(value)
        if type_string == "number_array":
            try:
                if "," in value:
                    return [
                        [Decimal(y) for y in x.split(",")] for x in value.split(";")
                    ]
                else:
                    return [Decimal(x) for x in value.split(";")]
            except (TypeError, ValueError, InvalidOperation):
                warn(
                    _(
                        'Non-numeric value "{}" found in number array column, returning as string array instead).'
                    ).format(value),
                    DataErrorWarning,
                )
        if "," in value:
            return [x.split(",") for x in value.split(";")]
        else:
            return value.split(";")
    elif type_string == "string":
        if type(value) == datetime.datetime:
            return timezone.localize(value).isoformat()
        return str(value)
    elif type_string == "date":
        if type(value) == datetime.datetime:
            return value.date().isoformat()
        return str(value)
    elif type_string == "":
        if type(value) == datetime.datetime:
            return timezone.localize(value).isoformat()
        if type(value) == float and int(value) == value:
            return int(value)
        return value if type(value) in [int] else str(value)
    else:
        raise ValueError('Unrecognised type: "{}"'.format(type_string))


def warnings_for_ignored_columns(v, extra_message):
    if isinstance(v, Cell):
        warn("Column {} has been ignored, {}".format(v.cell_location[3], extra_message))
    elif isinstance(v, dict):
        for x in v.values():
            warnings_for_ignored_columns(x, extra_message)
    elif isinstance(v, TemporaryDict):
        for x in v.to_list():
            warnings_for_ignored_columns(x, extra_message)
    else:
        raise ValueError()


def merge(base, mergee, debug_info=None):
    if not debug_info:
        debug_info = {}
    for key, v in mergee.items():
        if isinstance(v, Cell):
            value = v.cell_value
        else:
            value = v
        if key in base:
            if isinstance(value, TemporaryDict):
                if not isinstance(base[key], TemporaryDict):
                    warnings_for_ignored_columns(
                        v,
                        _(
                            "because it treats {} as an array, but another column does not"
                        ).format(key),
                    )
                    continue
                for temporarydict_key, temporarydict_value in value.items():
                    if temporarydict_key in base[key]:
                        merge(
                            base[key][temporarydict_key],
                            temporarydict_value,
                            debug_info,
                        )
                    else:
                        assert temporarydict_key not in base[key], _(
                            "Overwriting cell {} by mistake"
                        ).format(temporarydict_value)
                        base[key][temporarydict_key] = temporarydict_value
                for temporarydict_value in value.items_no_keyfield:
                    base[key].items_no_keyfield.append(temporarydict_value)
            elif isinstance(value, dict):
                if isinstance(base[key], dict):
                    merge(base[key], value, debug_info)
                else:
                    warnings_for_ignored_columns(
                        v,
                        _(
                            "because it treats {} as an object, but another column does not"
                        ).format(key),
                    )
            else:
                if not isinstance(base[key], Cell):
                    id_info = '{} "{}"'.format(
                        debug_info.get("id_name"),
                        debug_info.get(debug_info.get("id_name")),
                    )
                    if debug_info.get("root_id"):
                        id_info = (
                            '{} "{}", '.format(
                                debug_info.get("root_id"),
                                debug_info.get("root_id_or_none"),
                            )
                            + id_info
                        )
                    warnings_for_ignored_columns(
                        v, _("because another column treats it as an array or object")
                    )
                    continue
                base_value = base[key].cell_value
                if base_value != value:
                    id_info = '{} "{}"'.format(
                        debug_info.get("id_name"),
                        debug_info.get(debug_info.get("id_name")),
                    )
                    if debug_info.get("root_id"):
                        id_info = (
                            '{} "{}", '.format(
                                debug_info.get("root_id"),
                                debug_info.get("root_id_or_none"),
                            )
                            + id_info
                        )
                    warn(
                        _(
                            'You may have a duplicate Identifier: We couldn\'t merge these rows with the {}: field "{}" in sheet "{}": one cell has the value: "{}", the other cell has the value: "{}"'
                        ).format(
                            id_info,
                            key,
                            debug_info.get("sheet_name"),
                            base_value,
                            value,
                        ),
                        DataErrorWarning,
                    )
                else:
                    base[key].sub_cells.append(v)
        else:
            # This happens when a parent record finds the first a child record of a known type
            base[key] = v


class SpreadsheetInput(object):
    """
    Base class describing a spreadsheet input. Has stubs which are
    implemented via inheritance for particular types of spreadsheet (e.g. xlsx
    or csv).

    """

    def convert_dict_titles(self, dicts, title_lookup=None):
        """
        Replace titles with field names in the given list of dictionaries
        (``dicts``) using the titles lookup in the schema parser.

        """
        if self.parser:
            title_lookup = self.parser.title_lookup
        for d in dicts:
            if title_lookup:
                yield OrderedDict(
                    [(title_lookup.lookup_header(k), v) for k, v in d.items()]
                )
            else:
                yield d

    def __init__(
        self,
        input_name="",
        root_list_path="main",
        root_is_list=False,
        timezone_name="UTC",
        root_id="ocid",
        convert_titles=False,
        vertical_orientation=False,
        include_sheets=[],
        exclude_sheets=[],
        id_name="id",
        xml=False,
        base_configuration={},
        use_configuration=True,
    ):
        self.input_name = input_name
        self.root_list_path = root_list_path
        self.root_is_list = root_is_list
        self.sub_sheet_names = []
        self.timezone = pytz.timezone(timezone_name)
        self.root_id = root_id
        self.convert_titles = convert_titles
        self.id_name = id_name
        self.xml = xml
        self.parser = None
        self.vertical_orientation = vertical_orientation
        self.include_sheets = include_sheets
        self.exclude_sheets = exclude_sheets
        self.base_configuration = base_configuration or {}
        self.sheet_configuration = {}
        self.use_configuration = use_configuration

    def get_sub_sheets_lines(self):
        for sub_sheet_name in self.sub_sheet_names:
            if self.convert_titles:
                yield sub_sheet_name, self.convert_dict_titles(
                    self.get_sheet_lines(sub_sheet_name),
                    self.parser.sub_sheets[sub_sheet_name].title_lookup
                    if sub_sheet_name in self.parser.sub_sheets
                    else None,
                )
            else:
                yield sub_sheet_name, self.get_sheet_lines(sub_sheet_name)

    def configure_sheets(self):
        for sub_sheet_name in self.sub_sheet_names:
            self.sheet_configuration[sub_sheet_name] = parse_sheet_configuration(
                self.get_sheet_configuration(sub_sheet_name)
            )

    def get_sheet_configuration(self, sheet_name):
        return []

    def get_sheet_lines(self, sheet_name):
        raise NotImplementedError

    def get_sheet_headings(self, sheet_name):
        raise NotImplementedError

    def read_sheets(self):
        raise NotImplementedError

    def do_unflatten(self):
        main_sheet_by_ocid = OrderedDict()
        sheets = list(self.get_sub_sheets_lines())
        for i, sheet in enumerate(sheets):
            sheet_name, lines = sheet
            try:
                actual_headings = self.get_sheet_headings(sheet_name)
                # If sheet is empty or too many lines have been skipped
                if not actual_headings:
                    continue
                found = OrderedDict()
                last_col = len(actual_headings)
                # We want to ignore data in earlier columns, so we look
                # through the data backwards
                for i, actual_heading in enumerate(reversed(actual_headings)):
                    if actual_heading is None:
                        continue
                    if actual_heading in found:
                        found[actual_heading].append((last_col - i) - 1)
                    else:
                        found[actual_heading] = [i]
                for actual_heading in reversed(found):
                    if len(found[actual_heading]) > 1:
                        keeping = found[actual_heading][0]  # noqa
                        ignoring = found[actual_heading][1:]
                        ignoring.reverse()
                        if len(ignoring) >= 3:
                            warn(
                                (
                                    _(
                                        'Duplicate heading "{}" found, ignoring '
                                        'the data in columns {} and {} (sheet: "{}").'
                                    )
                                ).format(
                                    actual_heading,
                                    ", ".join(
                                        [
                                            _get_column_letter(x + 1)
                                            for x in ignoring[:-1]
                                        ]
                                    ),
                                    _get_column_letter(ignoring[-1] + 1),
                                    sheet_name,
                                ),
                                DataErrorWarning,
                            )
                        elif len(found[actual_heading]) == 3:
                            warn(
                                (
                                    _(
                                        'Duplicate heading "{}" found, ignoring '
                                        'the data in columns {} and {} (sheet: "{}").'
                                    )
                                ).format(
                                    actual_heading,
                                    _get_column_letter(ignoring[0] + 1),
                                    _get_column_letter(ignoring[1] + 1),
                                    sheet_name,
                                ),
                                DataErrorWarning,
                            )
                        else:
                            warn(
                                (
                                    _(
                                        'Duplicate heading "{}" found, ignoring '
                                        'the data in column {} (sheet: "{}").'
                                    )
                                ).format(
                                    actual_heading,
                                    _get_column_letter(ignoring[0] + 1),
                                    sheet_name,
                                ),
                                DataErrorWarning,
                            )
            except NotImplementedError:
                # The ListInput type used in the tests doesn't support getting headings.
                actual_headings = None
            for j, line in enumerate(lines):
                if all(x is None or x == "" for x in line.values()):
                    # if all(x == '' for x in line.values()):
                    continue
                root_id_or_none = line.get(self.root_id) if self.root_id else None
                cells = OrderedDict()
                for k, header in enumerate(line):
                    heading = actual_headings[k] if actual_headings else header
                    if self.vertical_orientation:
                        # This is misleading as it specifies the row number as the distance vertically
                        # and the horizontal 'letter' as a number.
                        # https://github.com/OpenDataServices/flatten-tool/issues/153
                        cells[header] = Cell(
                            line[header], (sheet_name, str(k + 1), j + 2, heading)
                        )
                    else:
                        cells[header] = Cell(
                            line[header],
                            (sheet_name, _get_column_letter(k + 1), j + 2, heading),
                        )
                unflattened = unflatten_main_with_parser(
                    self.parser, cells, self.timezone, self.xml, self.id_name
                )
                if root_id_or_none not in main_sheet_by_ocid:
                    main_sheet_by_ocid[root_id_or_none] = TemporaryDict(
                        self.id_name, xml=self.xml
                    )

                def inthere(unflattened, id_name):
                    if self.xml and not isinstance(unflattened.get(self.id_name), Cell):
                        # For an XML tag
                        return unflattened[id_name]["text()"].cell_value
                    else:
                        # For a JSON, or an XML attribute
                        return unflattened[id_name].cell_value

                if (
                    self.id_name in unflattened
                    and inthere(unflattened, self.id_name)
                    in main_sheet_by_ocid[root_id_or_none]
                ):
                    if self.xml and not isinstance(unflattened.get(self.id_name), Cell):
                        unflattened_id = unflattened.get(self.id_name)[
                            "text()"
                        ].cell_value
                    else:
                        unflattened_id = unflattened.get(self.id_name).cell_value
                    merge(
                        main_sheet_by_ocid[root_id_or_none][unflattened_id],
                        unflattened,
                        {
                            "sheet_name": sheet_name,
                            "root_id": self.root_id,
                            "root_id_or_none": root_id_or_none,
                            "id_name": self.id_name,
                            self.id_name: unflattened_id,
                        },
                    )
                else:
                    main_sheet_by_ocid[root_id_or_none].append(unflattened)
        temporarydicts_to_lists(main_sheet_by_ocid)
        return sum(main_sheet_by_ocid.values(), [])

    def unflatten(self):
        result = self.do_unflatten()
        result = extract_list_to_value(result)
        return result

    def fancy_unflatten(self, with_cell_source_map, with_heading_source_map):
        cell_tree = self.do_unflatten()
        result = extract_list_to_value(cell_tree)
        ordered_cell_source_map = None
        heading_source_map = None
        if with_cell_source_map or with_heading_source_map:
            cell_source_map = extract_list_to_error_path(
                [] if self.root_is_list else [self.root_list_path], cell_tree
            )
            ordered_items = sorted(cell_source_map.items())
            row_source_map = OrderedDict()
            heading_source_map = OrderedDict()
            for path, _unused in ordered_items:
                cells = cell_source_map[path]
                # Prepare row_source_map key
                key = "/".join(str(x) for x in path[:-1])
                if not key in row_source_map:
                    row_source_map[key] = []
                if with_heading_source_map:
                    # Prepare header_source_map key
                    header_path_parts = []
                    for x in path:
                        try:
                            int(x)
                        except:
                            header_path_parts.append(x)
                    header_path = "/".join(header_path_parts)
                    if header_path not in heading_source_map:
                        heading_source_map[header_path] = []
                # Populate the row and header source maps
                for cell in cells:
                    sheet, col, row, header = cell
                    if (sheet, row) not in row_source_map[key]:
                        row_source_map[key].append((sheet, row))
                    if with_heading_source_map:
                        if (sheet, header) not in heading_source_map[header_path]:
                            heading_source_map[header_path].append((sheet, header))
        if with_cell_source_map:
            ordered_cell_source_map = OrderedDict(
                ("/".join(str(x) for x in path), location)
                for path, location in ordered_items
            )
            for key in row_source_map:
                assert key not in ordered_cell_source_map, _(
                    "Row/cell collision: {}"
                ).format(key)
                ordered_cell_source_map[key] = row_source_map[key]
        return result, ordered_cell_source_map, heading_source_map


def extract_list_to_error_path(path, input):
    output = {}
    for i, item in enumerate(input):
        res = extract_dict_to_error_path(path + [i], item)
        for p in res:
            assert p not in output, _("Already have key {}").format(p)
            output[p] = res[p]
    return output


def extract_dict_to_error_path(path, input):
    output = {}
    for k in input:
        if isinstance(input[k], list):
            res = extract_list_to_error_path(path + [k], input[k])
            for p in res:
                assert p not in output, _("Already have key {}").format(p)
                output[p] = res[p]
        elif isinstance(input[k], dict):
            res = extract_dict_to_error_path(path + [k], input[k])
            for p in res:
                assert p not in output, _("Already have key {}").format(p)
                output[p] = res[p]
        elif isinstance(input[k], Cell):
            p = tuple(path + [k])
            assert p not in output, _("Already have key {}").format(p)
            output[p] = [input[k].cell_location]
            for sub_cell in input[k].sub_cells:
                assert sub_cell.cell_value == input[k].cell_value, _(
                    "Two sub-cells have different values: {}, {}"
                ).format(input[k].cell_value, sub_cell.cell_value)
                output[p].append(sub_cell.cell_location)
        else:
            raise Exception(
                _("Unexpected result type in the JSON cell tree: {}").format(input[k])
            )
    return output


def extract_list_to_value(input):
    output = []
    for item in input:
        output.append(extract_dict_to_value(item))
    return output


def extract_dict_to_value(input):
    output = OrderedDict()
    for k in input:
        if isinstance(input[k], list):
            output[k] = extract_list_to_value(input[k])
        elif isinstance(input[k], dict):
            output[k] = extract_dict_to_value(input[k])
        elif isinstance(input[k], Cell):
            output[k] = input[k].cell_value
        else:
            raise Exception(
                _("Unexpected result type in the JSON cell tree: {}").format(input[k])
            )
    return output


class CSVInput(SpreadsheetInput):
    encoding = "utf-8"

    def get_sheet_headings(self, sheet_name):
        sheet_configuration = self.sheet_configuration[self.sheet_names_map[sheet_name]]
        configuration_line = 1 if sheet_configuration else 0
        if not sheet_configuration:
            sheet_configuration = self.base_configuration
        if not self.use_configuration:
            sheet_configuration = {}
        skip_rows = sheet_configuration.get("skipRows", 0)
        if sheet_configuration.get("ignore"):
            # returning empty headers is a proxy for no data in the sheet.
            return []

        with open(
            os.path.join(self.input_name, sheet_name + ".csv"), encoding=self.encoding
        ) as main_sheet_file:
            r = csvreader(main_sheet_file)
            for num, row in enumerate(r):
                if num == (skip_rows + configuration_line):
                    return row

    def read_sheets(self):
        sheet_file_names = os.listdir(self.input_name)
        sheet_names = sorted(
            [fname[:-4] for fname in sheet_file_names if fname.endswith(".csv")]
        )
        if self.include_sheets:
            for sheet in list(sheet_names):
                if sheet not in self.include_sheets:
                    sheet_names.remove(sheet)
        for sheet in list(self.exclude_sheets) or []:
            try:
                sheet_names.remove(sheet)
            except ValueError:
                pass
        self.sub_sheet_names = sheet_names
        self.sheet_names_map = OrderedDict(
            (sheet_name, sheet_name) for sheet_name in sheet_names
        )
        self.configure_sheets()

    def generate_rows(self, dictreader, sheet_name):
        sheet_configuration = self.sheet_configuration[self.sheet_names_map[sheet_name]]
        configuration_line = 1 if sheet_configuration else 0
        if not sheet_configuration:
            sheet_configuration = self.base_configuration
        if not self.use_configuration:
            sheet_configuration = {}

        skip_rows = sheet_configuration.get("skipRows", 0)
        header_rows = sheet_configuration.get("headerRows", 1)
        for i in range(0, configuration_line + skip_rows):
            previous_row = next(dictreader.reader)  # noqa
        fieldnames = dictreader.fieldnames
        for i in range(0, header_rows - 1):
            next(dictreader.reader)
        for line in dictreader:
            yield OrderedDict((fieldname, line[fieldname]) for fieldname in fieldnames)

    def get_sheet_configuration(self, sheet_name):
        with open(
            os.path.join(self.input_name, sheet_name + ".csv"), encoding=self.encoding
        ) as main_sheet_file:
            r = csvreader(main_sheet_file)
            heading_row = next(r)
        if len(heading_row) > 0 and heading_row[0] == "#":
            return heading_row[1:]
        return []

    def get_sheet_lines(self, sheet_name):
        # Pass the encoding to the open function
        with open(
            os.path.join(self.input_name, sheet_name + ".csv"), encoding=self.encoding
        ) as main_sheet_file:
            dictreader = DictReader(main_sheet_file)
            for row in self.generate_rows(dictreader, sheet_name):
                yield row


class BadXLSXZipFile(BadZipFile):
    pass


class XLSXInput(SpreadsheetInput):
    def read_sheets(self):
        try:
            self.workbook = openpyxl.load_workbook(self.input_name, data_only=True)
        except BadZipFile as e:  # noqa
            # TODO when we have python3 only add 'from e' to show exception chain
            raise BadXLSXZipFile(
                _("The supplied file has extension .xlsx but isn't an XLSX file.")
            )

        self.sheet_names_map = OrderedDict(
            (sheet_name, sheet_name) for sheet_name in self.workbook.sheetnames
        )
        if self.include_sheets:
            for sheet in list(self.sheet_names_map):
                if sheet not in self.include_sheets:
                    self.sheet_names_map.pop(sheet)
        for sheet in self.exclude_sheets or []:
            self.sheet_names_map.pop(sheet, None)

        sheet_names = list(sheet for sheet in self.sheet_names_map.keys())
        self.sub_sheet_names = sheet_names
        self.configure_sheets()

    def get_sheet_headings(self, sheet_name):
        worksheet = self.workbook[self.sheet_names_map[sheet_name]]
        sheet_configuration = self.sheet_configuration[self.sheet_names_map[sheet_name]]
        configuration_line = 1 if sheet_configuration else 0
        if not sheet_configuration:
            sheet_configuration = self.base_configuration
        if not self.use_configuration:
            sheet_configuration = {}

        skip_rows = sheet_configuration.get("skipRows", 0)
        if sheet_configuration.get("ignore") or (
            sheet_configuration.get("hashcomments") and sheet_name.startswith("#")
        ):
            # returning empty headers is a proxy for no data in the sheet.
            return []

        if self.vertical_orientation:
            return [
                cell.value
                for cell in worksheet[_get_column_letter(skip_rows + 1)][
                    configuration_line:
                ]
            ]

        try:
            return [
                cell.value for cell in worksheet[skip_rows + configuration_line + 1]
            ]
        except IndexError:
            # If the heading line is after data in the spreadsheet. i.e when skipRows
            return []

    def get_sheet_configuration(self, sheet_name):
        worksheet = self.workbook[self.sheet_names_map[sheet_name]]
        if worksheet["A1"].value == "#":
            return [
                cell.value
                for num, cell in enumerate(worksheet[1])
                if num != 0 and cell.value
            ]
        else:
            return []

    def get_sheet_lines(self, sheet_name):
        sheet_configuration = self.sheet_configuration[self.sheet_names_map[sheet_name]]
        configuration_line = 1 if sheet_configuration else 0
        if not sheet_configuration:
            sheet_configuration = self.base_configuration
        if not self.use_configuration:
            sheet_configuration = {}

        skip_rows = sheet_configuration.get("skipRows", 0)
        header_rows = sheet_configuration.get("headerRows", 1)

        worksheet = self.workbook[self.sheet_names_map[sheet_name]]
        if self.vertical_orientation:
            header_row = worksheet[_get_column_letter(skip_rows + 1)]
            remaining_rows = worksheet.iter_cols(min_col=skip_rows + header_rows + 1)
            if configuration_line:
                header_row = header_row[1:]
                remaining_rows = worksheet.iter_cols(
                    min_col=skip_rows + header_rows + 1, min_row=2
                )
        else:
            header_row = worksheet[skip_rows + configuration_line + 1]
            remaining_rows = worksheet.iter_rows(
                min_row=skip_rows + configuration_line + header_rows + 1
            )

        coli_to_header = {}
        for i, header in enumerate(header_row):
            coli_to_header[i] = header.value

        for row in remaining_rows:
            output_row = OrderedDict()
            for i, x in enumerate(row):
                header = coli_to_header[i]
                value = x.value
                if not header:
                    # None means that the cell will be ignored
                    value = None
                elif (
                    sheet_configuration.get("hashcomments")
                    and isinstance(header, str)
                    and header.startswith("#")
                ):
                    # None means that the cell will be ignored
                    value = None
                output_row[header] = value
            yield output_row


class ODSInput(SpreadsheetInput):
    def read_sheets(self):
        self.workbook = ODSReader(self.input_name)
        self.sheet_names_map = self.workbook.SHEETS

        if self.include_sheets:
            for sheet in list(self.sheet_names_map):
                if sheet not in self.include_sheets:
                    self.sheet_names_map.pop(sheet)

        for sheet in self.exclude_sheets or []:
            self.sheet_names_map.pop(sheet, None)

        self.sub_sheet_names = self.sheet_names_map.keys()
        self.configure_sheets()

    def _resolve_sheet_configuration(self, sheet_name):
        sheet_configuration = self.sheet_configuration[sheet_name]
        if not self.use_configuration:
            return {"unused_config_line": True} if sheet_configuration else {}
        if not sheet_configuration:
            sheet_configuration = self.base_configuration
            sheet_configuration["base_configuration"] = True

        return sheet_configuration

    def get_sheet_headings(self, sheet_name):
        worksheet = self.sheet_names_map[sheet_name]

        sheet_configuration = self._resolve_sheet_configuration(sheet_name)
        configuration_line = (
            1
            if sheet_configuration and "base_configuration" not in sheet_configuration
            else 0
        )

        skip_rows = sheet_configuration.get("skipRows", 0)
        if sheet_configuration.get("ignore") or (
            sheet_configuration.get("hashcomments") and sheet_name.startswith("#")
        ):
            # returning empty headers is a proxy for no data in the sheet.
            return []

        if self.vertical_orientation:
            return [
                row[skip_rows]
                for row in worksheet[configuration_line:]
                if len(row) > skip_rows
            ]

        try:
            return [cell for cell in worksheet[skip_rows + configuration_line]]
        except IndexError:
            # If the heading line is after data in the spreadsheet. i.e when skipRows
            return []

    def get_sheet_configuration(self, sheet_name):
        # See if there are config properties in the spreadsheet
        # https://flatten-tool.readthedocs.io/en/latest/unflatten/#configuration-properties-skip-and-header-rows
        worksheet = self.sheet_names_map[sheet_name]

        try:
            # cell A1
            if worksheet[0][0] == "#":
                return worksheet[0]

        except IndexError:
            pass

        return []

    def get_sheet_lines(self, sheet_name):
        # This generator should yield an ordered dict in the format
        # see examples/simple/
        # yield OrderedDict([('a/b', '1'), ('a/c', '2'), ('d', '3')])
        # yield OrderedDict([('a/b', '4'), ('a/c', '5'), ('d', '6')])

        sheet_configuration = self._resolve_sheet_configuration(sheet_name)
        configuration_line = (
            1
            if sheet_configuration and "base_configuration" not in sheet_configuration
            else 0
        )

        skip_rows = sheet_configuration.get("skipRows", 0)
        header_rows = sheet_configuration.get("headerRows", 1)

        worksheet = self.sheet_names_map[sheet_name]
        if self.vertical_orientation:
            header_row = [
                row[skip_rows]
                for row in worksheet[configuration_line:]
                if len(row) > skip_rows
            ]
            longest_horizontal_row = max(
                len(row) for row in worksheet[configuration_line:]
            )
            remaining_rows = [
                [
                    row[i] if len(row) > i else None
                    for row in worksheet[configuration_line:]
                    if row
                ]
                for i in range(1, longest_horizontal_row)
            ]
        else:
            header_row = worksheet[skip_rows + configuration_line]
            remaining_rows = worksheet[(skip_rows + configuration_line + header_rows) :]

        coli_to_header = {}
        for i, header in enumerate(header_row):
            coli_to_header[i] = header

        for row in remaining_rows:
            output_row = OrderedDict()
            for i, x in enumerate(row):

                try:
                    header = coli_to_header[i]
                except KeyError:
                    continue
                value = x
                if not header:
                    # None means that the cell will be ignored
                    value = None
                elif sheet_configuration.get("hashcomments") and header.startswith("#"):
                    # None means that the cell will be ignored
                    value = None
                output_row[header] = value
            if output_row:
                if not all(value is None for value in output_row.values()):
                    yield output_row


FORMATS = {"xlsx": XLSXInput, "csv": CSVInput, "ods": ODSInput}


class ListAsDict(dict):
    pass


def list_as_dicts_to_temporary_dicts(unflattened, id_name, xml):
    for key, value in list(unflattened.items()):
        if isinstance(value, Cell):
            continue
        if hasattr(value, "items"):
            if not value:
                unflattened.pop(key)
            list_as_dicts_to_temporary_dicts(value, id_name, xml)
        if isinstance(value, ListAsDict):
            temporarydict = TemporaryDict(id_name, xml=xml)
            for index in sorted(value.keys()):
                temporarydict.append(value[index])
            unflattened[key] = temporarydict
    return unflattened


def unflatten_main_with_parser(parser, line, timezone, xml, id_name):
    unflattened = OrderedDict()
    for path, cell in line.items():
        # Skip blank cells
        if cell.cell_value is None or cell.cell_value == "":
            continue
        current_path = unflattened
        path_list = [item.rstrip("[]") for item in str(path).split("/")]
        for num, path_item in enumerate(path_list):
            if isint(path_item):
                if num == 0:
                    warn(
                        _(
                            'Column "{}" has been ignored because it is a number.'
                        ).format(path),
                        DataErrorWarning,
                    )
                continue
            current_type = None
            path_till_now = "/".join(
                [item for item in path_list[: num + 1] if not isint(item)]
            )
            if parser:
                current_type = parser.flattened.get(path_till_now)
            try:
                next_path_item = path_list[num + 1]
            except IndexError:
                next_path_item = ""

            # Quick solution to avoid casting of date as datetime in spreadsheet > xml
            if xml:
                if type(cell.cell_value) == datetime.datetime and not next_path_item:
                    if "datetime" not in str(path):
                        current_type = "date"

            ## Array
            list_index = -1
            if isint(next_path_item):
                if current_type and current_type != "array":
                    raise ValueError(
                        _(
                            "There is an array at '{}' when the schema says there should be a '{}'"
                        ).format(path_till_now, current_type)
                    )
                list_index = int(next_path_item)
                current_type = "array"

            if current_type == "array":
                list_as_dict = current_path.get(path_item)
                if list_as_dict is None:
                    list_as_dict = ListAsDict()
                    current_path[path_item] = list_as_dict
                elif type(list_as_dict) is not ListAsDict:
                    warn(
                        _(
                            "Column {} has been ignored, because it treats {} as an array, but another column does not."
                        ).format(path, path_till_now),
                        DataErrorWarning,
                    )
                    break
                new_path = list_as_dict.get(list_index)
                if new_path is None:
                    new_path = OrderedDict()
                    list_as_dict[list_index] = new_path
                current_path = new_path
                if not xml or num < len(path_list) - 2:
                    # In xml "arrays" can have text values, if they're the final element
                    # This corresponds to a tag with text, but also possibly attributes
                    continue

            ## Object
            if current_type == "object" or (not current_type and next_path_item):
                new_path = current_path.get(path_item)
                if new_path is None:
                    new_path = OrderedDict()
                    current_path[path_item] = new_path
                elif type(new_path) is ListAsDict or not hasattr(new_path, "items"):
                    warn(
                        _(
                            "Column {} has been ignored, because it treats {} as an object, but another column does not."
                        ).format(path, path_till_now),
                        DataErrorWarning,
                    )
                    break
                current_path = new_path
                continue
            if (
                current_type
                and current_type not in ["object", "array"]
                and next_path_item
            ):
                raise ValueError(
                    _(
                        "There is an object or list at '{}' but it should be an {}"
                    ).format(path_till_now, current_type)
                )

            ## Other Types
            current_path_value = current_path.get(path_item)
            if not xml and (
                type(current_path_value) is ListAsDict
                or hasattr(current_path_value, "items")
            ):
                #   ^
                # xml can have an object/array that also has a text value
                warn(
                    _(
                        "Column {} has been ignored, because another column treats it as an array or object"
                    ).format(path_till_now),
                    DataErrorWarning,
                )
                continue

            value = cell.cell_value
            if xml and current_type == "array":
                # In xml "arrays" can have text values, if they're the final element
                # However the type of the text value itself should not be "array",
                # as that would split the text on commas, which we don't want.
                # https://github.com/OpenDataServices/cove/issues/1030
                converted_value = convert_type("", value, timezone)
            else:
                converted_value = convert_type(current_type or "", value, timezone)
            cell.cell_value = converted_value
            if converted_value is not None and converted_value != "":
                if xml:
                    # For XML we want to support text and attributes at the
                    # same level, e.g.
                    # <my-element a="b">some text</my-element>
                    # which we represent in a dict as:
                    # {"@a":"b", "text()": "some text"}
                    # To ensure we can attach attributes everywhere, all
                    # element text must be added as a dict with a `text()` key.
                    if path_item.startswith("@"):
                        current_path[path_item] = cell
                    else:
                        if current_type == "array":
                            current_path["text()"] = cell
                        elif path_item not in current_path:
                            current_path[path_item] = {"text()": cell}
                        else:
                            current_path[path_item]["text()"] = cell
                else:
                    current_path[path_item] = cell

    unflattened = list_as_dicts_to_temporary_dicts(unflattened, id_name, xml)
    return unflattened


def path_search(
    nested_dict, path_list, id_fields=None, path=None, top=False, top_sheet=False
):
    if not path_list:
        return nested_dict

    id_fields = id_fields or {}
    parent_field = path_list[0]
    path = parent_field if path is None else path + "/" + parent_field

    if parent_field.endswith("[]") or top:
        if parent_field.endswith("[]"):
            parent_field = parent_field[:-2]
        if parent_field not in nested_dict:
            nested_dict[parent_field] = TemporaryDict(
                keyfield=id_name, top_sheet=top_sheet, xml=xml  # noqa
            )
        sub_sheet_id = id_fields.get(path + "/id")
        if sub_sheet_id not in nested_dict[parent_field]:
            nested_dict[parent_field][sub_sheet_id] = {}
        return path_search(
            nested_dict[parent_field][sub_sheet_id],
            path_list[1:],
            id_fields=id_fields,
            path=path,
            top_sheet=top_sheet,
        )
    else:
        if parent_field not in nested_dict:
            nested_dict[parent_field] = OrderedDict()
        return path_search(
            nested_dict[parent_field],
            path_list[1:],
            id_fields=id_fields,
            path=path,
            top_sheet=top_sheet,
        )


class TemporaryDict(UserDict):
    def __init__(self, keyfield, top_sheet=False, xml=False):
        self.keyfield = keyfield
        self.items_no_keyfield = []
        self.data = OrderedDict()
        self.top_sheet = top_sheet
        self.xml = xml

    def __repr__(self):
        return "TemporaryDict(keyfield={}, items_no_keyfield={}, data={})".format(
            repr(self.keyfield), repr(self.items_no_keyfield), repr(self.data)
        )

    def append(self, item):
        if self.keyfield in item:
            if self.xml:
                if isinstance(item[self.keyfield], Cell):
                    # For an XML attribute
                    key = item[self.keyfield].cell_value
                elif isinstance(item[self.keyfield]["text()"], Cell):
                    # For an XML tag
                    key = item[self.keyfield]["text()"].cell_value
                else:
                    key = item[self.keyfield]["text()"]
            else:
                if isinstance(item[self.keyfield], Cell):
                    key = item[self.keyfield].cell_value
                else:
                    key = item[self.keyfield]
            if key not in self.data:
                self.data[key] = item
            else:
                self.data[key].update(item)
        else:
            self.items_no_keyfield.append(item)

    def to_list(self):
        return list(self.data.values()) + self.items_no_keyfield


def temporarydicts_to_lists(nested_dict):
    """Recursively transforms TemporaryDicts to lists inplace."""
    for key, value in nested_dict.items():
        if isinstance(value, Cell):
            continue
        if hasattr(value, "to_list"):
            temporarydicts_to_lists(value)
            if hasattr(value, "items_no_keyfield"):
                for x in value.items_no_keyfield:
                    temporarydicts_to_lists(x)
            nested_dict[key] = value.to_list()
        elif hasattr(value, "items"):
            temporarydicts_to_lists(value)
