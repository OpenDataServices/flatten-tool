# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import os
import sys

import openpyxl
import pytest

from flattentool import output
from flattentool.ODSReader import ODSReader
from flattentool.sheet import Sheet


class MockParser(object):
    def __init__(self, main_sheet, sub_sheets):
        self.main_sheet = Sheet(main_sheet)
        self.sub_sheets = {k: Sheet(v) for k, v in sub_sheets.items()}


def test_spreadsheetouput_base_fails():
    """The base class should fail as it is missing functionality that child
    classes must implement"""

    spreadsheet_output = output.SpreadsheetOutput(parser=MockParser([], {}))
    with pytest.raises(NotImplementedError):
        spreadsheet_output.write_sheets()


def test_blank_sheets(tmpdir):
    for format_name, spreadsheet_output_class in output.FORMATS.items():
        spreadsheet_output = spreadsheet_output_class(
            parser=MockParser([], {}),
            main_sheet_name="release",
            output_name=os.path.join(
                tmpdir.strpath, "release" + output.FORMATS_SUFFIX[format_name]
            ),
        )
        spreadsheet_output.write_sheets()

    # Check XLSX is empty
    wb = openpyxl.load_workbook(tmpdir.join("release.xlsx").strpath)
    assert wb.sheetnames == ["release"]
    rows = list(wb["release"].rows)
    # openpyxl fixed this bug but earler versions of python are stuck with it.
    # remove when we no longer support 3.5
    if sys.version_info >= (3, 6, 0):
        assert len(rows) == 0

    # Check CSV is Empty
    assert tmpdir.join("release").listdir() == [
        tmpdir.join("release").join("release.csv")
    ]
    assert tmpdir.join("release", "release.csv").read().strip("\r\n") == ""

    # Check ODS is empty
    odswb = ODSReader(tmpdir.join("release.ods").strpath)
    ods_rows = odswb.getSheet("release")
    assert ods_rows == [[]]


def test_populated_header(tmpdir):
    for format_name, spreadsheet_output_class in output.FORMATS.items():
        subsheet = Sheet(root_id="ocid")
        subsheet.add_field("c")
        spreadsheet_output = spreadsheet_output_class(
            parser=MockParser(["a", "d"], {"b": subsheet}),
            main_sheet_name="release",
            output_name=os.path.join(
                tmpdir.strpath, "release" + output.FORMATS_SUFFIX[format_name]
            ),
        )
        spreadsheet_output.write_sheets()

    # Check XLSX
    wb = openpyxl.load_workbook(tmpdir.join("release.xlsx").strpath)
    assert wb.sheetnames == ["release", "b"]
    rows = list(wb["release"].rows)
    assert len(rows) == 1
    assert [x.value for x in rows[0]] == ["a", "d"]
    b_rows = list(wb["b"].rows)
    assert len(b_rows) == 1
    assert [x.value for x in b_rows[0]] == ["ocid", "c"]

    # Check CSV
    assert set(tmpdir.join("release").listdir()) == set(
        [
            tmpdir.join("release").join("release.csv"),
            tmpdir.join("release").join("b.csv"),
        ]
    )
    assert tmpdir.join("release", "release.csv").read().strip("\r\n") == "a,d"
    assert tmpdir.join("release", "b.csv").read().strip("\r\n") == "ocid,c"

    # Check ODS
    odswb = ODSReader(tmpdir.join("release.ods").strpath)
    ods_rows = odswb.getSheet("release")
    assert len(ods_rows) == 1
    assert [x for x in ods_rows[0]] == ["a", "d"]
    ods_b_rows = odswb.getSheet("b")
    assert len(ods_b_rows) == 1
    assert [x for x in ods_b_rows[0]] == ["ocid", "c"]


def test_empty_lines(tmpdir):
    subsheet = Sheet(root_id="ocid")
    subsheet.add_field("c")
    parser = MockParser(["a", "d"], {"b": subsheet})
    parser.main_sheet._lines = []
    for format_name, spreadsheet_output_class in output.FORMATS.items():
        spreadsheet_output = spreadsheet_output_class(
            parser=parser,
            main_sheet_name="release",
            output_name=os.path.join(
                tmpdir.strpath, "release" + output.FORMATS_SUFFIX[format_name]
            ),
        )
        spreadsheet_output.write_sheets()

    # Check XLSX
    wb = openpyxl.load_workbook(tmpdir.join("release.xlsx").strpath)
    assert wb.sheetnames == ["release", "b"]
    rows = list(wb["release"].rows)
    assert len(rows) == 1
    assert [x.value for x in rows[0]] == ["a", "d"]
    b_rows = list(wb["b"].rows)
    assert len(b_rows) == 1
    assert [x.value for x in b_rows[0]] == ["ocid", "c"]

    # Check CSV
    assert set(tmpdir.join("release").listdir()) == set(
        [
            tmpdir.join("release").join("release.csv"),
            tmpdir.join("release").join("b.csv"),
        ]
    )
    assert tmpdir.join("release", "release.csv").read().strip("\r\n") == "a,d"
    assert tmpdir.join("release", "b.csv").read().strip("\r\n") == "ocid,c"

    # Check ODS
    odswb = ODSReader(tmpdir.join("release.ods").strpath)
    ods_rows = odswb.getSheet("release")
    assert len(ods_rows) == 1
    assert [x for x in ods_rows[0]] == ["a", "d"]
    ods_b_rows = odswb.getSheet("b")
    assert len(ods_b_rows) == 1
    assert [x for x in ods_b_rows[0]] == ["ocid", "c"]


def test_populated_lines(tmpdir):
    subsheet = Sheet(root_id="ocid")
    subsheet.add_field("c")
    parser = MockParser(["a"], {})
    parser.main_sheet._lines = [{"a": "cell1"}, {"a": "cell2"}]
    subsheet._lines = [{"c": "cell3"}, {"c": "cell4"}]
    parser.sub_sheets["b"] = subsheet
    for format_name, spreadsheet_output_class in output.FORMATS.items():
        spreadsheet_output = spreadsheet_output_class(
            parser=parser,
            main_sheet_name="release",
            output_name=os.path.join(
                tmpdir.strpath, "release" + output.FORMATS_SUFFIX[format_name]
            ),
        )
        spreadsheet_output.write_sheets()

    # Check XLSX
    wb = openpyxl.load_workbook(tmpdir.join("release.xlsx").strpath)
    assert wb.sheetnames == ["release", "b"]
    rows = list(wb["release"].rows)
    assert len(rows) == 3
    assert [x.value for x in rows[0]] == ["a"]
    assert [x.value for x in rows[1]] == ["cell1"]
    assert [x.value for x in rows[2]] == ["cell2"]
    b_rows = list(wb["b"].rows)
    assert len(b_rows) == 3
    assert [x.value for x in b_rows[0]] == ["ocid", "c"]
    assert [x.value for x in b_rows[1]] == [None, "cell3"]
    assert [x.value for x in b_rows[2]] == [None, "cell4"]

    # Check CSV
    assert set(tmpdir.join("release").listdir()) == set(
        [
            tmpdir.join("release").join("release.csv"),
            tmpdir.join("release").join("b.csv"),
        ]
    )
    assert (
        tmpdir.join("release", "release.csv").read().strip("\r\n").replace("\r", "")
        == "a\ncell1\ncell2"
    )
    assert (
        tmpdir.join("release", "b.csv").read().strip("\r\n").replace("\r", "")
        == "ocid,c\n,cell3\n,cell4"
    )

    # Check ODS - currently broken test
    odswb = ODSReader(tmpdir.join("release.ods").strpath)
    ods_rows = odswb.getSheet("release")
    assert len(ods_rows) == 3
    assert [x for x in ods_rows[0]] == ["a"]
    assert [x for x in ods_rows[1]] == ["cell1"]
    assert [x for x in ods_rows[2]] == ["cell2"]
    ods_b_rows = odswb.getSheet("b")
    assert len(ods_b_rows) == 3
    assert [x for x in ods_b_rows[0]] == ["ocid", "c"]
    assert [x for x in ods_b_rows[1]] == [None, "cell3"]
    assert [x for x in ods_b_rows[2]] == [None, "cell4"]


def test_utf8(tmpdir):
    parser = MockParser(["é"], {})
    parser.main_sheet._lines = [{"é": "éαГ😼𝒞人"}, {"é": "cell2"}]
    for format_name, spreadsheet_output_class in output.FORMATS.items():
        spreadsheet_output = spreadsheet_output_class(
            parser=parser,
            main_sheet_name="release",
            output_name=os.path.join(
                tmpdir.strpath, "release" + output.FORMATS_SUFFIX[format_name]
            ),
        )
        spreadsheet_output.write_sheets()

    # Check XLSX
    wb = openpyxl.load_workbook(tmpdir.join("release.xlsx").strpath)
    assert wb.sheetnames == ["release"]
    rows = list(wb["release"].rows)
    assert len(rows) == 3
    assert [x.value for x in rows[0]] == ["é"]
    assert [x.value for x in rows[1]] == ["éαГ😼𝒞人"]
    assert [x.value for x in rows[2]] == ["cell2"]

    # Check CSV
    assert set(tmpdir.join("release").listdir()) == set(
        [
            tmpdir.join("release").join("release.csv"),
        ]
    )
    release_csv_text = tmpdir.join("release", "release.csv").read_text(encoding="utf-8")
    assert release_csv_text.strip("\r\n").replace("\r", "") == "é\néαГ😼𝒞人\ncell2"

    # Check ODS
    odswb = ODSReader(tmpdir.join("release.ods").strpath)
    ods_rows = odswb.getSheet("release")
    assert len(ods_rows) == 3
    assert [x for x in ods_rows[0]] == ["é"]
    assert [x for x in ods_rows[1]] == ["éαГ😼𝒞人"]
    assert [x for x in ods_rows[2]] == ["cell2"]
